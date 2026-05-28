"""
Excel-экспорт аналитики (PR D). Чистая функция → bytes, без сети/БД.

build_analytics_xlsx(data) собирает .xlsx из dict'а аналитики (тот же
формат что отдаёт /api/analytics для boss). Листы:
  • Сводка   — общие цифры (выручка, заказы, средний чек, тренд)
  • Клиенты  — топ клиентов по выручке
  • Менеджеры — топ менеджеров
  • Товары   — топ товаров + прибыль (где известна себестоимость)

Доставка — файлом в Telegram (см. /api/analytics/export). openpyxl
backward-compatible с Excel/Google Sheets/LibreOffice.
"""

from __future__ import annotations

import io
from typing import Any


def _autosize(ws, max_width: int = 50) -> None:
    """Подогнать ширину колонок под содержимое (грубо, по max длине)."""
    for col in ws.columns:
        length = 0
        letter = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is not None:
                length = max(length, len(str(v)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def build_analytics_xlsx(data: dict[str, Any]) -> bytes:
    """Собрать .xlsx из dict аналитики. Возвращает bytes (для send_document).

    `data` — структура из /api/analytics (boss): label/total/count/clients/
    avg_check/trend/top_clients/top_managers/top_products.
    Толерантна к отсутствующим ключам (пустые секции).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    # ─── Лист «Сводка» ───
    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Показатель", "Значение"])
    for c in ws[1]:
        c.font = bold
    ws.append(["Период", data.get("label", "")])
    ws.append(["Выручка", round(float(data.get("total", 0) or 0), 2)])
    ws.append(["Заказов", data.get("count", 0)])
    ws.append(["Клиентов", data.get("clients", 0)])
    ws.append(["Средний чек", round(float(data.get("avg_check", 0) or 0), 2)])
    ws.append(["Тренд, %", data.get("trend", 0)])
    _autosize(ws)

    # ─── Лист «Клиенты» ───
    ws_c = wb.create_sheet("Клиенты")
    ws_c.append(["Клиент", "Выручка", "Заказов"])
    for c in ws_c[1]:
        c.font = bold
    for row in data.get("top_clients", []) or []:
        ws_c.append(
            [row.get("name", "—"), round(float(row.get("revenue", 0) or 0), 2), row.get("count", 0)]
        )
    _autosize(ws_c)

    # ─── Лист «Менеджеры» ───
    ws_m = wb.create_sheet("Менеджеры")
    ws_m.append(["Менеджер", "Выручка", "Заказов"])
    for c in ws_m[1]:
        c.font = bold
    for row in data.get("top_managers", []) or []:
        ws_m.append(
            [row.get("name", "—"), round(float(row.get("revenue", 0) or 0), 2), row.get("count", 0)]
        )
    _autosize(ws_m)

    # ─── Лист «Товары» (+ прибыль где известна) ───
    ws_p = wb.create_sheet("Товары")
    ws_p.append(["Товар", "Кол-во", "Выручка", "Прибыль"])
    for c in ws_p[1]:
        c.font = bold
    for row in data.get("top_products", []) or []:
        profit = row.get("profit")
        ws_p.append(
            [
                row.get("name", "—"),
                row.get("qty", 0),
                round(float(row.get("sum", 0) or 0), 2),
                round(float(profit), 2) if profit is not None else "н/д",
            ]
        )
    _autosize(ws_p)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
