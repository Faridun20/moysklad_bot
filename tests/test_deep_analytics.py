"""
Тесты глубокой аналитики (PR D): per-client/manager агрегация, маржа,
произвольные даты, Excel-экспорт.

Граница с МС мокается aioresponses; БД настоящая (isolated_db) для cost.
"""

import asyncio
import importlib
import io
import re
from datetime import datetime

import pytest
from aioresponses import aioresponses

import services.moysklad as moysklad
from services.moysklad import MS_BASE, get_sales_stats, invalidate_ms_cache

_DEMAND_LIST_RE = re.compile(re.escape(f"{MS_BASE}/entity/demand") + r"\?.*")


def _positions_re(demand_id: str):
    return re.compile(re.escape(f"{MS_BASE}/entity/demand/{demand_id}/positions") + r".*")


def _reset_ms():
    moysklad._session = None
    invalidate_ms_cache()
    moysklad._circuit.record_success()


# ─── get_sales_stats: by_client + ms_id ──────────────────────────────────────


def test_sales_stats_includes_top_clients_and_ms_id():
    _reset_ms()
    since = datetime(2031, 4, 1)
    until = datetime(2031, 4, 2)
    shipments = {
        "rows": [
            {"meta": {"href": f"{MS_BASE}/entity/demand/D1"}, "sum": 100000, "agent": {"name": "Клиент-А"}},
            {"meta": {"href": f"{MS_BASE}/entity/demand/D2"}, "sum": 60000, "agent": {"name": "Клиент-А"}},
            {"meta": {"href": f"{MS_BASE}/entity/demand/D3"}, "sum": 40000, "agent": {"name": "Клиент-Б"}},
        ]
    }
    pos = [
        {
            "assortment": {"name": "Товар-1", "meta": {"href": f"{MS_BASE}/entity/product/PROD1"}},
            "quantity": 2,
            "price": 5000,
        }
    ]

    async def scenario():
        try:
            with aioresponses() as m:
                m.get(_DEMAND_LIST_RE, status=200, payload=shipments)
                for d in ("D1", "D2", "D3"):
                    m.get(_positions_re(d), status=200, payload=pos)
                return await get_sales_stats(since, until)
        finally:
            await moysklad.close_session()

    stats = asyncio.run(scenario())
    # by_client: Клиент-А = 160000, Клиент-Б = 40000
    clients = dict((n, d["sum"]) for n, d in stats["top_clients"])
    assert clients["Клиент-А"] == 160000
    assert clients["Клиент-Б"] == 40000
    assert stats["top_clients"][0][0] == "Клиент-А"  # сортировка по убыванию
    # ms_id извлечён в товарах
    prod = dict(stats["top_products"])
    assert prod["Товар-1"]["ms_id"] == "PROD1"


# ─── Excel-экспорт: build_analytics_xlsx ─────────────────────────────────────


def test_build_analytics_xlsx_valid_and_has_sheets():
    from services.excel_export import build_analytics_xlsx
    from openpyxl import load_workbook

    data = {
        "label": "Тест",
        "total": 1500.0,
        "count": 10,
        "clients": 5,
        "avg_check": 150.0,
        "trend": 12,
        "top_clients": [{"name": "Клиент-А", "revenue": 1000.0, "count": 6}],
        "top_managers": [{"name": "Менеджер-1", "revenue": 900.0, "count": 5}],
        "top_products": [
            {"name": "Товар-1", "qty": 3, "sum": 450.0, "profit": 150.0, "margin_known": True},
            {"name": "Товар-2", "qty": 1, "sum": 100.0, "margin_known": False},
        ],
    }
    xlsx = build_analytics_xlsx(data)
    assert isinstance(xlsx, bytes) and len(xlsx) > 0

    wb = load_workbook(io.BytesIO(xlsx))
    assert set(wb.sheetnames) == {"Сводка", "Клиенты", "Менеджеры", "Товары"}
    # Сводка содержит выручку
    summary_vals = [c.value for row in wb["Сводка"].iter_rows() for c in row]
    assert 1500.0 in summary_vals
    # Товары: профит известен у одного, «н/д» у другого
    prod_vals = [c.value for row in wb["Товары"].iter_rows() for c in row]
    assert 150.0 in prod_vals
    assert "н/д" in prod_vals


def test_build_analytics_xlsx_handles_empty_sections():
    from services.excel_export import build_analytics_xlsx
    from openpyxl import load_workbook

    xlsx = build_analytics_xlsx({"label": "Пусто"})
    wb = load_workbook(io.BytesIO(xlsx))
    # Листы есть даже без данных (только заголовки)
    assert "Товары" in wb.sheetnames


# ─── E2E: /api/analytics (даты) + /api/analytics/export ──────────────────────


@pytest.fixture
def client_env(isolated_db, monkeypatch):
    import services.roles as roles
    import webapp.server as server

    importlib.reload(roles)
    db = isolated_db
    boss_id, mgr_id = 100, 200
    db.set_role(boss_id, "boss", "Boss", "boss")
    db.set_role(mgr_id, "mgr", "Manager", "manager")

    monkeypatch.setattr(
        server,
        "verify_init_data",
        lambda init_data: {"id": int(init_data), "first_name": "U", "username": "u"},
    )
    return server, db, {"boss": boss_id, "mgr": mgr_id}


def test_resolve_period_custom_dates_validation():
    """_resolve_analytics_period: плохие даты → 400, валидные → диапазон."""
    import webapp.server as server
    from fastapi import HTTPException

    now = datetime(2031, 6, 1)
    # Невалидный формат
    with pytest.raises(HTTPException) as e1:
        server._resolve_analytics_period({"since": "не-дата", "until": "2031-05-01"}, now)
    assert e1.value.status_code == 400
    # until ≤ since
    with pytest.raises(HTTPException) as e2:
        server._resolve_analytics_period({"since": "2031-05-10", "until": "2031-05-01"}, now)
    assert e2.value.status_code == 400
    # Слишком большой диапазон
    with pytest.raises(HTTPException) as e3:
        server._resolve_analytics_period({"since": "2020-01-01", "until": "2031-01-01"}, now)
    assert e3.value.status_code == 400
    # Валидный
    since, until, prev, label = server._resolve_analytics_period(
        {"since": "2031-05-01", "until": "2031-05-15"}, now
    )
    assert since == datetime(2031, 5, 1)
    assert until == datetime(2031, 5, 15)
    # WP-20: until с фронта эксклюзивный → в метке показываем ВЫБРАННЫЙ конец
    # (until − 1 день), а не сам эксклюзивный until.
    assert label == "2031-05-01 — 2031-05-14"


def test_resolve_period_preset_returns_now_as_until():
    import webapp.server as server

    now = datetime(2031, 6, 1)
    since, until, prev, label = server._resolve_analytics_period({"period": "week"}, now)
    assert until == now
    assert label == "Неделя"


def test_analytics_export_forbidden_for_manager(client_env, monkeypatch):
    from fastapi.testclient import TestClient

    server, _db, ids = client_env
    client = TestClient(server.app)
    resp = client.post("/api/analytics/export", json={"initData": str(ids["mgr"]), "period": "week"})
    assert resp.status_code == 403


def test_manager_performance_splits_currencies(isolated_db):
    """get_manager_performance: выручка/долг РАЗДЕЛЬНО по валютам (топ-менеджеров
    в company-аналитике больше не складывает USD+EUR в одно число с «$»)."""
    import datetime as dt

    db = isolated_db
    db.set_role(5, "m", "Менеджер", "manager")
    for cur, qty, price in [("USD", 2, 100.0), ("EUR", 1, 50.0)]:
        oid = db.create_order(5, "Менеджер", "")
        db.update_order_agent(oid, "A", "Client")
        db.update_order_currency(oid, cur)
        db.add_order_item(oid, "Товар", "", qty, "шт", price)
        db.update_order_status(oid, "shipped")
    since = (dt.datetime.now() - dt.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
    until = (dt.datetime.now() + dt.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
    perf = asyncio.run(db.get_manager_performance(since, until))
    m = next(x for x in perf if x["user_id"] == 5)
    rev = {x["currency"]: x["amount"] for x in m["revenue_by_currency"]}
    assert rev.get("USD") == 200.0
    assert rev.get("EUR") == 50.0


def test_personal_analytics_splits_currencies(client_env):
    """Личная аналитика менеджера НЕ складывает разные валюты: выручка
    возвращается списком {currency, total} по каждой валюте отдельно."""
    from fastapi.testclient import TestClient

    server, db, ids = client_env
    mgr = ids["mgr"]

    def mk(cur, qty, price):
        oid = db.create_order(mgr, "M", "")
        db.update_order_agent(oid, "A" + cur, "Client " + cur)
        db.update_order_currency(oid, cur)
        db.add_order_item(oid, "Товар", "", qty, "шт", price)
        db.update_order_status(oid, "shipped")
        return oid

    mk("USD", 2, 100.0)      # 200 USD
    mk("UZS", 3, 50000.0)    # 150 000 UZS

    client = TestClient(server.app)
    resp = client.post("/api/analytics", json={"initData": str(mgr), "period": "month"})
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["scope"] == "personal"
    rev = {r["currency"]: r["total"] for r in body["revenue"]}
    assert rev.get("USD") == 200.0
    assert rev.get("UZS") == 150000.0
    assert len(body["revenue"]) == 2  # не схлопнуто в одно число


def test_company_analytics_degrades_when_ms_down(client_env, monkeypatch):
    """MS недоступен (401/сеть/5xx) → /api/analytics НЕ 500, а 200 с
    ms_unavailable=True. Регресс на «Не удалось загрузить: Unexpected token
    'I', "Internal S"… is not valid JSON» (500-боди — не JSON)."""
    from fastapi.testclient import TestClient

    import services.moysklad as moysklad

    async def _boom(*a, **k):
        raise RuntimeError("MS 401 Unauthorized")

    # _company_analytics_payload делает `from services.moysklad import …`
    # на каждый вызов → патч атрибутов модуля перехватывается.
    monkeypatch.setattr(moysklad, "get_sales_stats", _boom)
    monkeypatch.setattr(moysklad, "get_shipments", _boom)

    server, _db, ids = client_env
    client = TestClient(server.app)
    resp = client.post("/api/analytics", json={"initData": str(ids["boss"]), "period": "week"})
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["ms_unavailable"] is True
    assert body["total"] == 0
    assert body["scope"] == "company"


def test_analytics_export_sends_document_for_boss(client_env, monkeypatch):
    from fastapi.testclient import TestClient

    server, _db, ids = client_env

    # Мокаем расчёт аналитики (не ходим в МС) и notify-bot.
    async def fake_payload(since, until, prev_since, label):
        return {
            "label": label,
            "total": 100.0,
            "count": 1,
            "clients": 1,
            "avg_check": 100.0,
            "trend": 0,
            "top_products": [],
            "top_clients": [],
            "top_managers": [],
        }

    sent = {}

    class _FakeBot:
        async def send_document(self, chat_id, document, caption=None):
            sent["chat_id"] = chat_id
            sent["filename"] = getattr(document, "filename", None)

    async def fake_get_bot():
        return _FakeBot()

    monkeypatch.setattr(server, "_company_analytics_payload", fake_payload)
    monkeypatch.setattr(server, "get_notify_bot", fake_get_bot)

    client = TestClient(server.app)
    resp = client.post(
        "/api/analytics/export", json={"initData": str(ids["boss"]), "period": "month"}
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["sent"] is True
    assert sent["chat_id"] == ids["boss"]
    assert sent["filename"].endswith(".xlsx")
